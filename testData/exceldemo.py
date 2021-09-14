import openpyxl
book = openpyxl.load_workbook("E:\\My Selenium & Python\\SampleDemoUsingExcel.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Manoj S K"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A4'].value)
# Now using for loop to print all excel values in console
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column+1):
            #print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)