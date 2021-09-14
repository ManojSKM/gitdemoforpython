import openpyxl


class HomePageMultipleTestData:
    test_homepage_data = [{"Name":"Manoj", "Email":"srimanojskm@gmail.com", "PWD":"samsungj7max12!@", "Gender":"Male"}, {"Name":"Prabala", "Email":"prabaskmano@gmail.com", "PWD":"samsungj7max12!@", "Gender":"Female"}, {"Name":"Krishnamoorthy", "Email":"krishnamoorthy@gmail.com", "PWD":"samsungj7max12!@", "Gender":"Male"}]

    @staticmethod
    def gettestdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("E:\\My Selenium & Python\\SampleDemoUsingExcel.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        #print(Dict)
        return [Dict]
